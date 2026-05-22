from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
SOURCE_FILES = [
    Path("/Users/mustafa/Desktop/assets/данные о товарах/Контент Sanuzelshop водонагреватель.xlsx"),
    Path("/Users/mustafa/Desktop/assets/данные о товарах/Контент Sanuzelshop Зеркало.xlsx"),
    Path("/Users/mustafa/Desktop/assets/данные о товарах/Контент Sanuzelshop унитаз.xlsx"),
]
OUTPUT_DIR = ROOT_DIR / "frontend" / "data" / "products"

CATEGORY_ORDER = ["toilets", "smart-toilets", "sinks", "mirrors", "water-heaters"]

STRAPI_CATEGORIES = {
    "Унитаз": {
        "key": "toilets",
        "strapiId": 1,
        "name": "Унитазы",
        "englishName": "Toilets",
        "slug": "unitazy",
    },
    "Унитаз умный": {
        "key": "smart-toilets",
        "strapiId": 4,
        "name": "Умные унитазы",
        "englishName": "Smart Toilets",
        "slug": "umnye-unitazy",
    },
    "Раковина": {
        "key": "sinks",
        "strapiId": 6,
        "name": "Раковины",
        "englishName": "Sinks",
        "slug": "rakoviny",
    },
    "Зеркало": {
        "key": "mirrors",
        "strapiId": 10,
        "name": "Зеркала",
        "englishName": "Mirrors",
        "slug": "zerkala",
    },
    "Водонагреватель": {
        "key": "water-heaters",
        "strapiId": 8,
        "name": "Водонагреватели",
        "englishName": "Water Heaters",
        "slug": "vodonagrevateli",
    },
}

CORE_HEADERS = {
    "Наименование",
    "Артикулы В ПРАЙСЕ",
    "Артикул",
    "Категория",
    "Производитель",
    "Дополнительное описание",
    "URL",
}

ATTRIBUTE_MAP = {
    "Вес": ("weight", "Вес", None),
    "Тип изделия": ("productType", "Тип изделия", None),
    "Цвет": ("color", "Цвет", None),
    "Ширина": ("widthMm", "Ширина", "mm"),
    "Высота": ("heightMm", "Высота", "mm"),
    "Глубина": ("depthMm", "Глубина", "mm"),
    "Длина": ("lengthMm", "Длина", "mm"),
    "Тип установки": ("installationType", "Тип установки", None),
    "Мощность": ("powerW", "Мощность", "W"),
    "Страна происхождения": ("countryOfOrigin", "Страна происхождения", None),
    "Гарантия": ("warranty", "Гарантия", None),
    "Материал корпуса": ("bodyMaterial", "Материал корпуса", None),
    "Покрытие корпуса": ("bodyFinish", "Покрытие корпуса", None),
    "Материал фасада": ("facadeMaterial", "Материал фасада", None),
    "Способ монтажа": ("mountingMethod", "Способ монтажа", None),
    "Тип лампы": ("lampType", "Тип лампы", None),
    "Цвет подсветки": ("lightingColor", "Цвет подсветки", None),
    "Монтаж": ("mounting", "Монтаж", None),
    "Направление выпуска": ("outletDirection", "Направление выпуска", None),
    "Вид смывающего потока": ("flushFlowType", "Вид смывающего потока", None),
    "Цвет сиденья": ("seatColor", "Цвет сиденья", None),
    "Цвет фурнитуры": ("hardwareColor", "Цвет фурнитуры", None),
    "Поверхность": ("surface", "Поверхность", None),
    "Материал": ("material", "Материал", None),
}

IMAGE_ROLES = {
    "Основное изображение": "main",
    "Увеличенное изображение": "detail",
    "Увеличенное изображение 1": "detail",
    "Увеличенное изображение 2": "detail",
    "Чертеж": "drawing",
    "Вид спереди": "front",
    "Вид сзади": "back",
    "Вид снизу": "bottom",
    "Вид сбоку": "side",
    "Готовый узел": "assembled",
    "В интерьере": "interior",
    "В закрытом виде": "closed",
    "Вид с открытой крышкой": "open-lid",
}

VIDEO_HEADERS = {"Видео"}
PRICE_TERMS = ("цена", "стоимость", "price", "retail", "ррц")
EMPTY_MARKERS = {"", ".", "-", "—", "нет", "n/a", "null", "none"}

CYRILLIC_TO_LATIN = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "kh",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)

SKU_HOMOGLYPHS = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "Х": "X",
        "а": "a",
        "е": "e",
        "о": "o",
        "р": "p",
        "с": "c",
        "х": "x",
        "у": "y",
    }
)


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        normalized = clean_text(value)
        return normalized is None or normalized.lower() in EMPTY_MARKERS
    return False


def clean_text(value: Any, multiline: bool = False) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace("\u00a0", " ").replace("_x000D_", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.strip().strip("\ufeff")
    while len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        text = text[1:-1].strip()
    if multiline:
        lines: list[str] = []
        previous_blank = False
        for line in text.split("\n"):
            cleaned = re.sub(r"[ \t]+", " ", line).strip()
            if not cleaned:
                if lines and not previous_blank:
                    lines.append("")
                previous_blank = True
                continue
            lines.append(cleaned)
            previous_blank = False
        while lines and lines[-1] == "":
            lines.pop()
        text = "\n".join(lines)
    else:
        text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in EMPTY_MARKERS:
        return None
    return text or None


def normalize_sku(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text.translate(SKU_HOMOGLYPHS)


def normalize_scalar(value: Any) -> str | int | float | bool | None:
    if is_empty(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    return clean_text(value)


def parse_price(value: Any) -> int | float | None:
    if is_empty(value):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"[^\d,.-]", "", text).replace(" ", "")
    if not text:
        return None
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        price = float(text)
    except ValueError:
        return None
    return int(price) if price.is_integer() else price


def slugify(value: str) -> str:
    lowered = value.lower().translate(CYRILLIC_TO_LATIN)
    slug = re.sub(r"[^a-z0-9]+", "-", lowered)
    return re.sub(r"-{2,}", "-", slug).strip("-")


def unique_value(base: str, used: set[str]) -> str:
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def remove_brand_prefix(name: str, brand: str | None) -> str:
    if not brand:
        return name
    return re.sub(rf"^{re.escape(brand)}\.\s*", "", name, flags=re.IGNORECASE).strip()


def extract_model(name: str, sku: str | None) -> str | None:
    if sku:
        return sku
    match = re.search(r"\bЗеркало\s+([A-Za-zА-Яа-я0-9-]+)", name, flags=re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None


def normalize_description(description: str, product_name: str, sku: str | None) -> str:
    lines = []
    for line in description.split("\n"):
        if line.startswith("Название:"):
            lines.append(f"Название: {product_name}")
            continue
        if sku and line.startswith("Артикул:"):
            lines.append(f"Артикул: {sku}")
            continue
        lines.append(line)
    return "\n".join(lines).strip()


def compact_dimensions(row: dict[str, Any]) -> str | None:
    dimensions = []
    for key in ("Ширина", "Высота", "Глубина", "Длина"):
        value = normalize_scalar(row.get(key))
        if isinstance(value, (int, float)):
            dimensions.append(str(value))
    return "x".join(dimensions) if dimensions else None


def build_images(row: dict[str, Any], product_name: str) -> list[dict[str, Any]]:
    images: list[dict[str, Any]] = []
    used_urls: set[str] = set()
    for header, role in IMAGE_ROLES.items():
        value = clean_text(row.get(header))
        if not value or not value.startswith(("http://", "https://")):
            continue
        if value in used_urls:
            continue
        used_urls.add(value)
        images.append(
            {
                "url": value,
                "role": role,
                "label": header,
                "alt": product_name,
            }
        )
    return images


def build_videos(row: dict[str, Any]) -> list[str]:
    videos = []
    for header in VIDEO_HEADERS:
        value = clean_text(row.get(header))
        if value and value.startswith(("http://", "https://")):
            videos.append(value)
    return videos


def build_attributes(row: dict[str, Any]) -> list[dict[str, Any]]:
    attributes: list[dict[str, Any]] = []
    for header, value in row.items():
        if header in CORE_HEADERS or header in IMAGE_ROLES or header in VIDEO_HEADERS:
            continue
        if any(term in header.lower() for term in PRICE_TERMS):
            continue
        normalized = normalize_scalar(value)
        if normalized is None:
            continue
        key, label, unit = ATTRIBUTE_MAP.get(
            header,
            (slugify(header).replace("-", "_"), header, None),
        )
        item: dict[str, Any] = {
            "key": key,
            "label": label,
            "value": normalized,
        }
        if unit:
            item["unit"] = unit
        attributes.append(item)
    return attributes


def detect_price(row: dict[str, Any]) -> int | float | None:
    for header, value in row.items():
        if any(term in header.lower() for term in PRICE_TERMS):
            price = parse_price(value)
            if price is not None:
                return price
    return None


def row_to_product(row: dict[str, Any]) -> dict[str, Any] | None:
    raw_name = clean_text(row.get("Наименование"))
    category_name = clean_text(row.get("Категория"))
    if not raw_name or not category_name:
        return None
    if category_name not in STRAPI_CATEGORIES:
        raise ValueError(f"Unknown category {category_name!r} for product {raw_name!r}")

    category = STRAPI_CATEGORIES[category_name]
    brand = clean_text(row.get("Производитель"))
    sku = normalize_sku(row.get("Артикул"))
    product_name = remove_brand_prefix(raw_name, brand)
    if sku:
        raw_sku = clean_text(row.get("Артикул"))
        if raw_sku and raw_sku != sku:
            product_name = product_name.replace(raw_sku, sku)
    model = extract_model(product_name, sku)
    description = clean_text(row.get("Дополнительное описание"), multiline=True) or ""
    if sku:
        raw_sku = clean_text(row.get("Артикул"))
        if raw_sku and raw_sku != sku:
            description = description.replace(raw_sku, sku)
    description = normalize_description(description, product_name, sku)

    dimensions = compact_dimensions(row)
    slug_base_parts = [brand, product_name]
    if not sku and dimensions:
        slug_base_parts.append(dimensions)
    slug_base = slugify(" ".join(part for part in slug_base_parts if part))

    identity_parts = [category["key"], model or product_name]
    if not sku and dimensions:
        identity_parts.append(dimensions)
    id_base = "product-" + slugify("-".join(identity_parts))

    price = detect_price(row)
    return {
        "idBase": id_base,
        "slugBase": slug_base,
        "dedupeKey": (
            category["key"],
            sku.lower() if sku else slugify(product_name),
            dimensions or "",
            clean_text(row.get("Цвет")) or "",
        ),
        "product": {
            "id": "",
            "slug": "",
            "sku": sku,
            "name": product_name,
            "brand": brand,
            "model": model,
            "category": category,
            "price": price,
            "currency": "RUB" if price is not None else None,
            "description": description,
            "images": build_images(row, product_name),
            "videos": build_videos(row),
            "attributes": build_attributes(row),
        },
    }


def read_products() -> tuple[list[dict[str, Any]], int]:
    raw_count = 0
    product_wrappers: list[dict[str, Any]] = []
    for source_file in SOURCE_FILES:
        workbook = load_workbook(source_file, data_only=True)
        worksheet = workbook["Товары"]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                header: row_values[index]
                for index, header in enumerate(headers)
                if header
            }
            if all(is_empty(value) for value in row.values()):
                continue
            raw_count += 1
            wrapper = row_to_product(row)
            if wrapper:
                product_wrappers.append(wrapper)
    return product_wrappers, raw_count


def normalize_products() -> tuple[list[dict[str, Any]], dict[str, int]]:
    wrappers, raw_count = read_products()
    products_by_key: dict[tuple[Any, ...], dict[str, Any]] = {}
    for wrapper in wrappers:
        key = wrapper["dedupeKey"]
        product = wrapper["product"]
        if key not in products_by_key:
            products_by_key[key] = wrapper
            continue

        existing = products_by_key[key]["product"]
        existing_urls = {image["url"] for image in existing["images"]}
        for image in product["images"]:
            if image["url"] not in existing_urls:
                existing["images"].append(image)
                existing_urls.add(image["url"])
        existing_videos = set(existing["videos"])
        for video in product["videos"]:
            if video not in existing_videos:
                existing["videos"].append(video)
                existing_videos.add(video)

    used_ids: set[str] = set()
    used_slugs: set[str] = set()
    products: list[dict[str, Any]] = []
    for wrapper in products_by_key.values():
        product = wrapper["product"]
        product["id"] = unique_value(wrapper["idBase"], used_ids)
        product["slug"] = unique_value(wrapper["slugBase"], used_slugs)
        products.append(product)

    category_index = {key: index for index, key in enumerate(CATEGORY_ORDER)}
    products.sort(
        key=lambda product: (
            category_index.get(product["category"]["key"], 999),
            product["brand"] or "",
            product["name"],
            product["sku"] or "",
        )
    )
    return products, {
        "sourceRows": raw_count,
        "products": len(products),
        "duplicatesRemoved": raw_count - len(products),
    }


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    products, stats = normalize_products()
    categories = [STRAPI_CATEGORIES[key] for key in STRAPI_CATEGORIES]
    categories.sort(key=lambda category: CATEGORY_ORDER.index(category["key"]))

    global_payload = {
        "schemaVersion": 1,
        "categories": categories,
        "products": products,
    }
    write_json(OUTPUT_DIR / "products.json", global_payload)

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        grouped[product["category"]["key"]].append(product)

    for category in categories:
        payload = {
            "schemaVersion": 1,
            "category": category,
            "products": grouped.get(category["key"], []),
        }
        write_json(OUTPUT_DIR / f"{category['key']}.json", payload)

    print(json.dumps(stats, ensure_ascii=False, indent=2))
    for category in categories:
        print(f"{category['key']}: {len(grouped.get(category['key'], []))}")


if __name__ == "__main__":
    main()
